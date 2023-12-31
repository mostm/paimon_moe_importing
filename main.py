import json
from typing import List, Dict, Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime


def load_banner_history():
    with open('banner_history.json', 'r') as f:
        banners = json.load(f)
    return banners


def convert_to_hg_format():
    return


def load_banner_history_xlsx():
    workbook: Workbook = openpyxl.load_workbook('paimonmoe_wish_history.xlsx')
    workbook[0]: Worksheet
    pass


def find_banner(banners: List[Dict], wish: Dict) -> Optional[Dict]:
    # Convert the time string in the wish to a UNIX timestamp
    wish_time = datetime.strptime(wish['time'], '%Y-%m-%d %H:%M:%S').timestamp()

    split_banners = {}
    for banner in banners:
        if not split_banners.get(banner['type']):
            split_banners[banner['type']] = []
        split_banners[banner['type']].append(banner)

    # sort banners
    for banner_type in split_banners:
        split_banners[banner_type] = sorted(split_banners[banner_type], key=lambda x: x['when'], reverse=False)

    if wish['gacha_type'] not in split_banners.keys():
        return None

    proper_banners = split_banners[wish['gacha_type']]
    for i, banner in enumerate(proper_banners):
        # If gacha_type matches
        # print(f'banner: {banner} | {datetime.fromtimestamp(banner["when"])}')
        # If wish_time is within the banner's time range
        after_start = banner['when'] <= wish_time
        before_end = i == len(proper_banners) - 1 or wish_time < proper_banners[i + 1]['when']
        # print(f'after_start: {after_start}, before_end: {before_end}')
        if after_start and before_end:
            return banner
    return None


def fill_history(wb: Workbook, banners: list[dict], wish_history: list[dict]):
    # fill headers
    for sheet_name in wb.sheetnames[:-2]:
        sheet = wb[sheet_name]
        sheet.append(['Type', 'Name', 'Time', '⭐', 'Pity', '#Roll', 'Group', 'Banner', 'Part'])

    pity_counters = {}
    roll_count = {}

    for wish in wish_history:

        banner = find_banner(banners, wish)

        if banner is None and wish['gacha_type'] in ['301', '302', '400']:
            print(f'Could not find banner for wish: {wish}')
            continue
        else:
            if wish['gacha_type'] == "100":
                banner = {"name": "Beginners' Wish", "type": "100"}
            elif wish['gacha_type'] == "200":
                banner = {"name": "Wanderlust Invocation", "type": "200"}

        sheet_name = 'Beginners\' Wish'
        if wish['gacha_type'] == '100':
            sheet_name = 'Beginners\' Wish'
        elif wish['gacha_type'] == '200':
            sheet_name = 'Standard'
        elif wish['gacha_type'] == '301':
            sheet_name = 'Character Event'
        elif wish['gacha_type'] == '302':
            sheet_name = 'Weapon Event'
        elif wish['gacha_type'] == '400':
            sheet_name = 'Character Event'

        sheet = wb[sheet_name]
        entry = []

        # Type
        entry.append(wish['item_type'])

        # Name
        entry.append(wish['name'])

        # Time
        entry.append(wish['time'])

        # ⭐
        entry.append(wish['rank_type'])

        # Pity
        pity_type = "301" if wish['gacha_type'] == '400' else wish['gacha_type']

        if not pity_counters.get(pity_type):
            pity_counters[pity_type] = {"4": 0, "5": 0}

        pity = 1
        if wish['rank_type'] in ['5', '4']:
            pity = pity_counters[pity_type][wish['rank_type']]
        entry.append(str(pity))

        if wish['rank_type'] in ['5', '4']:
            pity_counters[pity_type][wish['rank_type']] = 0
        else:
            pity_counters[pity_type]['5'] += 1
            pity_counters[pity_type]['4'] += 1
        print(f'{pity_type} | {wish["gacha_type"]}')

        # #Roll
        roll_count[banner['name']] = roll_count.get(banner['name'], 0) + 1
        entry.append(roll_count[banner['name']])

        # Group
        entry.append(roll_count[banner['name']])  # Example: if 10-pull, they all belong to one group. Resets for each banner.

        # Banner
        entry.append(banner['name'])

        # Part
        entry.append("")

        sheet.append(entry)

    return


def generate_history():
    with open('banner_history.json', 'r') as f:
        banners = json.load(f)

    with open('genshin_wish_history.json', 'r') as f:
        genshin_wish_history = json.load(f)

    wb = openpyxl.Workbook()
    wb.create_sheet('Character Event')
    wb.create_sheet('Weapon Event')
    wb.create_sheet('Standard')
    wb.create_sheet('Beginners\' Wish')
    wb.create_sheet('Banner List')
    wb.create_sheet('Information')

    fill_history(wb, banners, reversed(genshin_wish_history['history']))

    # Information sheet needs to have "Paimon.moe Wish History Export" in A1 or the improt fails according to paimon.moe's source code
    wb['Information']['A1'] = 'Paimon.moe Wish History Export'

    wb.save('generated_history.xlsx')

    return


def main():
    generate_history()


if __name__ == '__main__':
    main()
